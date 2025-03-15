import requests
from data_handler import get_handler
from DrissionPage import ChromiumOptions, ChromiumPage
# from selenium import webdriver
import pandas as pd
# from openpyxl import Workbook, load_workbook
import os
from lxml import etree
# from bs4 import BeautifulSoup
import re
import time
import random
import json
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from spider import data_handler
from state_logger import *
from typing import Tuple, Optional
# from deepl_translate import translate_text_deepl

class Spider:
    def __init__(self):
        # co = ChromiumOptions()
        # co.set_browser_path(r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe")
        # co.use_system_user_path()
        # self.browser = ChromiumPage(co)

        self.url = "https://sh.5i5j.com/ershoufang/"
        self.cookies = {
            'SECKEY_ABVK': 'u/tTJNif2iYZ6hNp9pzyWnRCNlBczmqqAHrQtNWyTwI%3D',
            'BMAP_SECKEY': 'hqNqhvAjUam_yWrrahaCl7axbAo1DDcmF9fIVVWQCHjXDIj9dGyKfdxuhxshwDUmRbyNV9yzgCc37ydFg0BVDkAyhe1ZEfbkMCr24dVSAao-9j7kBQjTWfEE8HsUTJJFxLMQdp-YD9OZjOcCvi1wFC6t3VoKQmwMR6Ayhb2RTxnAhiUL_lTkcEqez2uKnINC',
            'ser_cookie_id': 'A53CC0BC-3598-216C-548E-65F9514BB84B',
            'sensorsdata2015jssdkchannel': '%7B%22prop%22%3A%7B%22_sa_channel_landing_url%22%3A%22%22%7D%7D',
            'smidV2': '2025012316020288ac7b9c44e79062312e87d37c08416e00d1ea797b1370ec0',
            'gr_user_id': '92d160ed-db39-4152-a026-92a071404e2e',
            'sensorsdata2015jssdkcross': '%7B%22distinct_id%22%3A%22194922e698f6fc-0ee02f325b71908-4c657b58-2764800-194922e6990f9d%22%2C%22first_id%22%3A%22%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E7%9B%B4%E6%8E%A5%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC_%E7%9B%B4%E6%8E%A5%E6%89%93%E5%BC%80%22%2C%22%24latest_referrer%22%3A%22%22%2C%22_latest_wx_ad_click_id%22%3A%22%22%2C%22_latest_wx_ad_hash_key%22%3A%22%22%2C%22_latest_wx_ad_callbacks%22%3A%22%22%7D%2C%22identities%22%3A%22eyIkaWRlbnRpdHlfY29va2llX2lkIjoiMTk0OTIyZTY5OGY2ZmMtMGVlMDJmMzI1YjcxOTA4LTRjNjU3YjU4LTI3NjQ4MDAtMTk0OTIyZTY5OTBmOWQifQ%3D%3D%22%2C%22history_login_id%22%3A%7B%22name%22%3A%22%22%2C%22value%22%3A%22%22%7D%2C%22%24device_id%22%3A%22194922e698f6fc-0ee02f325b71908-4c657b58-2764800-194922e6990f9d%22%7D',
            'ershoufang_BROWSES': '825090979%2C824673028%2C501273599%2C825098628',
            'PHPSESSID': 'evqq8tv42j6fna0ku99i10kkl6',
            'Hm_lvt_94ed3d23572054a86ed341d64b267ec6': '1737619324,1739093266,1739864416',
            'HMF_CI': '9c27b4825d1cf597e342ae0610ed51c22db9e880a3421037a9102303f9e0ac6267517adae53579999b482bc73e41e4413481c34302b40e46b072e0a86d0a91cd80',
            'HMACCOUNT': 'D8D5D3657EA1B9FA',
            'Hm_lvt_cf8004879455b04c74d33aa164379a1d': '1737619303,1739093266,1739864442',
            'domain': 'sh',
            'HMY_JC': '460ce0f8a5cb4fd5a8cc33f2693112330f81d6bf15908497101586417d806eaefe,',
            '8fcfcf2bd7c58141_gr_session_id': '1965859f-9a9b-483f-880d-11336928d5a9',
            'Hm_lpvt_94ed3d23572054a86ed341d64b267ec6': '1740145199',
            'Hm_lpvt_cf8004879455b04c74d33aa164379a1d': '1740145199',
            'C3VK': '049571',
        }  # self._get_cookies(True)
        self.headers = {
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36 Edg/121.0.0.0',
        }
        self.proxies = {  # Clash Verge
            "http": "http://127.0.0.1:7890",
            "https": "http://127.0.0.1:7890",
        }

        # service = webdriver.EdgeService('C:\\Users\\11627\\OneDrive\\CS\\PL\\Python\\Scraping\\drivers\\msedgedriver.exe')
        # options = webdriver.EdgeOptions()
        # # 无头浏览器设置
        # # options.add_argument('--headless')
        # # options.add_argument('--disable-gpu')
        # # 反反selenium爬虫设置
        # # options.add_experimental_option('excludeSwitches', ['enable-automation'])
        # # options.add_experimental_option('useAutomationExtension', False)
        # # options.add_experimental_option('detach', True)  # 运行结束不关闭浏览器进程

        # self.browser = webdriver.Edge(options=options, service=service)
        # self.browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        #     "source": "Object.defineProperty(navigator, 'webdriver', {get:()=>undefined})"
        # })

        self.data_handler = get_handler('csv', repo_name='sh_resale_house')
        
        # 需要将列定义移动到 DataHandler 初始化参数中（或提前定义在类属性）
        self._init_columns()  # todo: 列名在什么地方定义？

    def _init_columns(self):
        """初始化数据列定义"""
        self.columns = [
            "标题", "房型", "方向", "总价", "单价", 
            "面积", "楼层", "建造年份", "建筑类型",
            "行政区", "街道", "小区", "地址", "图片", 
            "标签", "网址"
        ]
        # 如果使用需要列名的处理器（如 CSV/Excel），需要传递列信息
        if hasattr(self.data_handler, 'repo'):
            if isinstance(self.data_handler.repo, pd.DataFrame) and self.data_handler.repo.empty:
                self.data_handler.repo = pd.DataFrame(columns=self.columns)

    def get_processed_data(self, raw_data):
        response = requests.post(
            "http://127.0.0.1:3000",
            json={
                "rawData": raw_data,
            }
        )
        processed_data = response.text
        return processed_data

    def crawl_list_page(self, page, index):
        response = requests.get(
            self.url + f"n{page}/",
            headers=self.headers
        )

        # print(response.text)

        if "oopopdwskl" in response.text:
            sleep_time = 3600
            print(f"触发风控，等待 {sleep_time} 秒后继续")
            time.sleep(sleep_time)

        # 使用 Xpath 解析 HTML
        tree = etree.HTML(response.text)
        detail_url_list = tree.xpath('/html/body/div[7]/div[1]/div[2]/ul/li/div[2]/h3/a/@href')[:30]
        for detail_url_index, detail_url in enumerate(detail_url_list):
            if detail_url_index < index:
                continue

            house_id = detail_url.split("/")[-1].split(".")[0]
            tag = tree.xpath(f'/html/body/div[7]/div[1]/div[2]/ul/li[{index + 1}]/div[2]/div[2]/span/text()')
            data = [*self.crawl_detail_page(house_id), tag, self.url + detail_url.split("/")[-1] + "/"]
            # print(data)
            self.data_handler.append_data(data)
            self.data_handler.save_data()
            print("detail_url_index", detail_url_index, "next_index", detail_url_index + 1)
            save_state(page, detail_url_index + 1)
            sleep_time = random.randint(10, 20)
            print(f"已爬取 {detail_url} 并保存数据，等待 {sleep_time} 秒后继续")
            time.sleep(sleep_time)

    def crawl_detail_page(self, house_id):
        max_retries = 3  # 最大重试次数
        attempt = 0

        while attempt < max_retries:
            try:
                url = "https://appapi.5i5j.com/vr/9/houseinfo"
                data = {
                    'cityid': '9',
                    'houseid': house_id,
                    'type': '1',
                    'brokerId': '',
                    # 'bid': '565924',
                    'equipment': '',
                    'from': '',
                    'platform': '1',
                    'vrcomplanyid': '4',
                }

                response = requests.post(
                    url,
                    headers=self.headers,
                    data=data
                )

                house_info = response.json()['data']["houseInfoNew"]
                title = house_info["title"]
                layout = house_info["layout"]
                head_ing = house_info["heading"]
                price = house_info["price"]
                unit_price = house_info["unitprice"]
                area = house_info["area"]
                floor = house_info["floor"]
                build_year = house_info["buildyear"]
                building_type = house_info["buildingtype"]
                sq = house_info["sq"]
                community_name = house_info["communityName"]
                address = house_info["address"]
                imgs_url = house_info["imgs"]

                lat = house_info["lat"]
                lng = house_info["lng"]

                url = "https://sh.5i5j.com/periphery"
                params = {
                    'query': '地铁站',
                    'location': f'{lat}, {lng}',
                    'radius': '3000',
                    'scope': '2',
                    'page_num': '0',
                }
                response = requests.get(
                    url,
                    params=params,
                    headers={
                        **self.headers,
                        "referer":f"{self.url}/{house_id}.html"
                    }
                )
                # print(response.json())
                distinct = response.json()['data'][0]['area']

                return title, layout, head_ing, price, unit_price, area, floor, build_year, building_type, distinct, sq, community_name, address, imgs_url
            except Exception as e:
                attempt += 1
                print(f"第 {attempt} 次尝试失败: {e}")
                if attempt >= max_retries:
                    print(url)
                    print("所有重试均失败。")
                    raise e
                
                sleep_time = random.randint(10, 20)
                print(f"等待 {sleep_time} 秒后进行第 {attempt + 1} 次重试...")
                time.sleep(sleep_time)

    def crawl_image(self, url, folder_path):
        """
        从指定 URL 下载图片并保存到本地路径。

        参数:
            url (str): 图片的 URL。
            save_path (str): 图片的保存路径（包括文件名）。

        返回:
            bool: 下载是否成功。
        """
        try:
            # 确保保存路径的目录存在
            filename = url.split('/')[-1]
            save_path = folder_path + "/" + filename
            if os.path.exists(save_path):
                print(f"图片已存在: {save_path}")
                return save_path
            os.makedirs(folder_path, exist_ok=True)
            
            # 发送 HTTP 请求获取图片数据
            response = requests.get(url, stream=True)
            response.raise_for_status()  # 检查请求是否成功

            # 将图片数据写入文件
            with open(save_path, "wb") as file:
                for chunk in response.iter_content(chunk_size=8192):
                    file.write(chunk)

            print(f"图片已成功保存到: {save_path}")
            return save_path
        except requests.exceptions.RequestException as e:
            print(f"下载图片失败: {e}")
            return False
        except Exception as e:
            print(f"保存图片失败: {e}")
            return False

    # def insert_image(self, image_paths, col_idx, idx):
    #     """
    #     将多张图片合并为一张图片后插入到指定列的单元格中。如果只有一张图片，则只调整大小并插入。

    #     Args:
    #         image_paths (list): 图片路径列表。
    #         col_idx (int): 目标列的索引（从 1 开始）。
    #     """
    #     # 如果只有一张图片，则直接调整大小并插入
    #     if len(image_paths) == 1:
    #         img_path = image_paths[0]
    #         with PILImage.open(img_path) as img:
    #             # 调整图片大小
    #             original_width, original_height = img.size
    #             aspect_ratio = original_width / original_height
    #             target_height = 100
    #             target_width = int(target_height * aspect_ratio)
    #             resized_img = img.resize((target_width, target_height), PILImage.Resampling.LANCZOS)
    #     else:
    #         # 合并多张图片
    #         resized_img = self.merge_images(image_paths)

    #     # 获取最后一行的指定列单元格
    #     last_row = self.excel.active.max_row  # 最后一行
    #     target_cell = self.excel.active.cell(row=last_row, column=col_idx)

    #     # 保存调整大小或合并后的图片到临时文件
    #     temp_image_path = f"temp_image_{idx}.png"
    #     resized_img.save(temp_image_path)  # 确保保存为 temp_image.png

    #     # 插入图片到 Excel
    #     img = Image(temp_image_path)
    #     self.excel.active.add_image(img, target_cell.coordinate)  # 将图片插入到目标单元格

    #     # 调整行高和列宽
    #     image_height = resized_img.height  # 图片高度（像素）
    #     image_width = resized_img.width  # 图片宽度（像素）

    #     # 调整行高
    #     row_height = image_height / 1.3328  # 将像素转换为 Excel 行高单位（点）
    #     self.excel.active.row_dimensions[last_row].height = row_height

    #     # 调整列宽
    #     column_width = image_width / 7.5  # 将像素转换为 Excel 列宽单位（字符宽度）
    #     if column_width > self.max_column_width:  # 如果列宽超过最大列宽，则调整
    #         self.excel.active.column_dimensions[
    #             self.excel.active.cell(row=1, column=col_idx).column_letter].width = column_width
    #         self.max_column_width = column_width

    #     # # 删除临时图片文件
    #     # os.remove(temp_image_path)

    # def merge_images(self, image_paths, spacing=10):
    #     """
    #     将多张图片合并为一张图片。

    #     Args:
    #         image_paths (list): 图片路径列表。
    #         spacing (int): 图片之间的间距。

    #     Returns:
    #         PILImage.Image: 合并后的图片。
    #     """
    #     images = []  # 存储处理后的图片
    #     total_width = 0  # 合并后图片的总宽度

    #     # 遍历图片路径，等比缩放图片高度为 100
    #     for path in image_paths:
    #         with PILImage.open(path) as img:
    #             original_width, original_height = img.size
    #             aspect_ratio = original_width / original_height
    #             target_height = 100
    #             target_width = int(target_height * aspect_ratio)
    #             resized_img = img.resize((target_width, target_height), PILImage.Resampling.LANCZOS)
    #             images.append(resized_img)
    #             total_width += target_width

    #     # 计算合并后图片的总宽度（包括间距）
    #     total_width += spacing * (len(image_paths) - 1)

    #     # 创建合并后的图片
    #     merged_image = PILImage.new("RGB", (total_width, 100), (255, 255, 255))  # 白色背景
    #     x_offset = 0
    #     for img in images:
    #         merged_image.paste(img, (x_offset, 0))
    #         x_offset += img.width + spacing

    #     return merged_image
            
    def save_issued_page_and_exit(self, response_text):
        with open("issued_page.html", "w", encoding="utf-8") as f:
            f.write(response_text)
        exit(0)

if __name__ == "__main__":
    # 需要更新主程序的处理器初始化方式
    spider = Spider()
    
    try:
        page_origin = 1

        # 检查是否有未完成的状态
        state = load_state()
        if state:
            start_page = state["next_page"]
            start_index = state["next_index"]
        else:
            start_page = page_origin
            start_index = 0
        page = start_page
        end_page = 177
        if not page_origin:
            end_page -= 1
        while True:
            if page > end_page:
                break
            spider.crawl_list_page(page, start_index)

            sleep_time = random.randint(10, 20)
            print(f"已爬取第 {page if page_origin else page + 1} 页并保存数据，等待 {sleep_time} 秒后继续")
            page += 1
            save_state(page, 0)
            time.sleep(sleep_time)
        
        clear_state()
        print("爬取完成~~~！")

        # # 默认使用Excel处理（与之前完全兼容）
        # spider = Spider()

        # # 显式指定Excel（支持自定义参数）
        # spider = Spider(db_params={
        #     'custom_filename': 'custom_data.xlsx'  # 可扩展自定义文件名
        # })

        # # 使用PostgreSQL
        # spider = Spider(
        #     storage_type='sqlalchemy',
        #     url='postgresql://user:pass@localhost:5432/mydb'
        # )

        # # 使用MongoDB
        # spider = Spider(
        #     storage_type='mongodb',
        #     uri='mongodb://localhost:27017/',
        #     dbname='housing_data'
        # )
    finally:
        spider.data_handler.close_repository()